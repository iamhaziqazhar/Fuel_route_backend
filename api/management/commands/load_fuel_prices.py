from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Iterator

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from api.models import FuelPrice


REQUIRED_FIELDS = ('city', 'state', 'latitude', 'longitude')
PRICE_FIELDS = ('regular_price', 'mid_grade_price', 'premium_price', 'diesel_price')
ALL_FIELDS = REQUIRED_FIELDS + PRICE_FIELDS


class Command(BaseCommand):
    help = 'Load fuel station prices from a CSV or Excel file into the FuelPrice table.'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the CSV or .xlsx file.')
        parser.add_argument(
            '--sheet',
            type=str,
            default=None,
            help='Excel sheet name (defaults to the active/first sheet).',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Parse and validate without writing to the database.',
        )

    def handle(self, *args, **options):
        file_path = Path(options['file_path']).expanduser().resolve()
        if not file_path.is_file():
            raise CommandError(f'File not found: {file_path}')

        suffix = file_path.suffix.lower()
        if suffix == '.csv':
            rows = self._read_csv(file_path)
        elif suffix in {'.xlsx', '.xlsm'}:
            rows = self._read_excel(file_path, sheet=options['sheet'])
        else:
            raise CommandError(
                f'Unsupported file type: {suffix}. Expected .csv, .xlsx, or .xlsm.'
            )

        created = updated = skipped = 0
        errors: list[str] = []

        with transaction.atomic():
            for line_no, raw in enumerate(rows, start=2):
                cleaned, err = self._clean_row(raw)
                if err:
                    skipped += 1
                    errors.append(f'  row {line_no}: {err}')
                    continue

                if options['dry_run']:
                    created += 1
                    continue

                lat = cleaned.pop('latitude')
                lng = cleaned.pop('longitude')
                _, was_created = FuelPrice.objects.update_or_create(
                    latitude=lat,
                    longitude=lng,
                    defaults=cleaned,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            if options['dry_run']:
                transaction.set_rollback(True)

        verb = 'Would load' if options['dry_run'] else 'Loaded'
        self.stdout.write(self.style.SUCCESS(
            f'{verb} fuel prices from {file_path.name}: '
            f'{created} created, {updated} updated, {skipped} skipped.'
        ))
        if errors:
            self.stdout.write(self.style.WARNING(
                f'\nSkipped {len(errors)} row(s):'
            ))
            for line in errors[:20]:
                self.stdout.write(line)
            if len(errors) > 20:
                self.stdout.write(f'  ... and {len(errors) - 20} more')

    def _read_csv(self, path: Path) -> Iterator[dict]:
        with path.open(newline='', encoding='utf-8-sig') as fh:
            reader = csv.DictReader(fh)
            self._check_headers(reader.fieldnames or [])
            yield from reader

    def _read_excel(self, path: Path, sheet: str | None) -> Iterator[dict]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                'Reading .xlsx files requires openpyxl. Install with: '
                'pip install openpyxl'
            ) from exc

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return
        headers = [str(h).strip() if h is not None else '' for h in header_row]
        self._check_headers(headers)

        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == '' for c in row):
                continue
            yield {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}

    def _check_headers(self, headers: Iterable[str]) -> None:
        present = {h.strip().lower() for h in headers if h}
        missing = [f for f in REQUIRED_FIELDS if f not in present]
        if missing:
            raise CommandError(
                f'Missing required column(s): {", ".join(missing)}. '
                f'Expected at least: {", ".join(REQUIRED_FIELDS)}.'
            )

    def _clean_row(self, raw: dict) -> tuple[dict, str | None]:
        row = {(k or '').strip().lower(): v for k, v in raw.items()}

        cleaned: dict = {}

        for field in ('city', 'state'):
            value = row.get(field)
            if value is None or str(value).strip() == '':
                return {}, f'missing {field}'
            cleaned[field] = str(value).strip()

        for field in ('latitude', 'longitude'):
            try:
                cleaned[field] = self._to_decimal(row.get(field), required=True)
            except ValueError as exc:
                return {}, f'{field}: {exc}'

        for field in PRICE_FIELDS:
            try:
                cleaned[field] = self._to_decimal(row.get(field), required=False)
            except ValueError as exc:
                return {}, f'{field}: {exc}'

        if cleaned['regular_price'] is None and cleaned['mid_grade_price'] is None:
            return {}, 'no regular_price or mid_grade_price (cannot price 10 MPG vehicle)'

        return cleaned, None

    @staticmethod
    def _to_decimal(value, *, required: bool) -> Decimal | None:
        if value is None or (isinstance(value, str) and value.strip() == ''):
            if required:
                raise ValueError('missing value')
            return None
        try:
            return Decimal(str(value).strip().replace('$', '').replace(',', ''))
        except (InvalidOperation, ValueError):
            raise ValueError(f'not a number: {value!r}')
